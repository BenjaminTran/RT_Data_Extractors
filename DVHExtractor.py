#import gc
import numpy as np
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl import worksheet
from openpyxl.utils import get_column_letter, get_column_interval
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import xlFunctions as xlF
import os
import conf
import pickle

def main():
	data = []
	directory = 'data/tandemANDring/converted/'

	master_df = pd.DataFrame(columns=(['Patient ID'] + [variable for variable in conf.DVH_VARIABLES] + conf.DVH_OTHER_VARIABLES))

	file_list = os.listdir(directory)
	file_list.sort()

	for zdx,filename in enumerate(file_list):
		print(f"filename is: {filename}")
		f = os.path.join(directory, filename)

		wb = load_workbook(filename=f, data_only = True)
		sheet = wb.active

		cell_container = xlF.xl_container(sheet)

		# tuple of the cells with value in the str argument
		structure_tuple = xlF.cell_string(cell_container,'Structure')
		# tuple of the cells with value in the str argument
		DVH_tuple = xlF.cell_string(cell_container,'Dose [cGy]   Relative dose [%] Ratio of Total Structure Volume [%]')

		# Obtain the fraction number
		fraction_num_cell = xlF.cell_string(cell_container, 'Plan')[0]
		fraction_num_row = fraction_num_cell.row
		fraction_num_col = get_column_letter(fraction_num_cell.column + 1)
		#plan_fraction = sheet[fraction_num_col + str(fraction_num_row)].value.split('_')
		#try:
		#	fraction_char_list = list(plan_fraction[1])
		#except:
		#	plan_fraction = sheet[fraction_num_col + str(fraction_num_row)].value.split('-')
		fraction_char_list = list(sheet[fraction_num_col + str(fraction_num_row)].value)
		Fraction = -1
		number_string = ''
		for char in fraction_char_list:
			if char.isdigit():
				number_string+=char
		try:
			Fraction = int(number_string)
		except ValueError:
			print(f"Fraction number was not found in plan. Plan text is {fraction_char_list}")
		# Obtain the plan
		#Plan = plan_fraction[0]

		# Obtain DVH variables of interest
		## This will be a list of lists
		dvh_variables = []
		for variable in conf.DVH_VARIABLES:
			dvh_variables.append(xlF.DVH_field_extractor(cell_container,sheet,variable))
		patient_id = xlF.DVH_field_extractor(cell_container,sheet,'Patient ID')


		# Create list of structure names
		structure_list = []
		for item in structure_tuple:
		    coordinate = coordinate_from_string(item.coordinate)
		    row = coordinate[1]
		    col = get_column_letter(item.column + 1)
		    if conf.DEBUG:
		    	print(f"{col} + {row}")
		    structure_list.append(sheet[col+str(row)].value.split('_')[0])

		for idx,structure in enumerate(structure_tuple):
		    print(f"THE INDEX IS: {idx}")
		    # Begin filling list to insert into top level dataframe
		    insertion_list = [value_list[idx] for value_list in dvh_variables]
		    DF_list = patient_id + insertion_list + [Fraction,structure_list[idx]]
		    print(f"length of DF_list: {len(DF_list)}")

		    # Obtain DVH data
		    coordinate = coordinate_from_string(structure.coordinate)
		    ## dict to create DVH dataframe
		    extracted_data_dict = {'Dose': [],
		    					'Relative Dose': [],
		    					'Ratio of Total Structure Volume': []}
		    DVH_row_start = DVH_tuple[idx].row + 1
		    ## If not the last structure then last row of DVH data for the structure is as follows
		    if idx < len(structure_tuple) - 1:
		    	DVH_row_end = structure_tuple[idx+1].row
		    else:
		    	DVH_row_end = sheet.max_row
		    for row_num in range(DVH_row_start,DVH_row_end):
		    	value_string = sheet[coordinate[0] + str(row_num)].value
		    	try:
		    		data_vector = value_string.split()
		    	except:
		    		print(row_num)
		    	for jdx,key in enumerate(extracted_data_dict.keys()):
		    		#extracted_data_dict[key].append(float(data_vector[jdx]))
		    		extracted_data_dict[key].append(float(data_vector[jdx]))
		    # Create the DVH dataframe
		    dvh_df = pd.DataFrame(data=extracted_data_dict)
		    # Convert values to float32 to reduce memory usage by 50% 
		    for col in dvh_df.columns:
		    	dvh_df[col] = dvh_df[col].astype('float32')
		    # Insert DVH dataframe into list for insertion into top level dataframe
		    DF_list.append(dvh_df)
		    # Append to top level dataframe
		    master_df.loc[len(master_df)] = DF_list
		wb.close()
	# Save as pickle file to retrieve later
	master_df.to_pickle("./test.pkl")

if __name__ == '__main__':
	main()