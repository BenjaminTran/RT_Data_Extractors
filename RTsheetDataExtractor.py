import openpyxl
from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl import worksheet
from openpyxl.utils import get_column_letter, get_column_interval
import xlFunctions as xlF
import conf
import os

def main():
	data = []
	directory = 'data/RTSheets/'

	# Setup workbook to write data
	wb_new = openpyxl.Workbook()
	wb_new_sheet = wb_new.active
	# add desired variables to extract as headers. Nb. 'Fraction' and 'Identifier' must be added manually
	wb_new_sheet.append(conf.VARIABLES)
	wb_new_sheet.insert_cols(1)
	wb_new_sheet['A1'] = 'Fraction'
	wb_new_sheet.insert_cols(1)
	wb_new_sheet['A1'] = 'Identifier'

	for filename in os.listdir(directory):
		f = os.path.join(directory, filename)

		wb = load_workbook(filename=f, data_only = True)
		sheet = wb.active

		cell_container = xlF.xl_container(sheet)

		for idx in range(0,xlF.number_of_fractions(cell_container)):
		    data.append(xlF.patient_identifier(cell_container, sheet))
		    data.append(idx+1)
		    for i in range(0,len(conf.VARIABLES)):
		        data.append(xlF.data_extractor(cell_container,sheet,conf.VARIABLES[i],idx+1))
		    wb_new_sheet.append(data)
		    data.clear()

	# save the new workbook
	wb_new.save('Output.xlsx')

if __name__ == "__main__":
	main()
