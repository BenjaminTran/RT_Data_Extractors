### This file uses features (f strings) that require python 3.6

import datetime as dt
import sys
from openpyxl.utils import get_column_letter
import warnings
import conf
import numbers

def xl_container(sheet) -> list:
	"""
	Returns a list of all cells with a value in the sheet. Used to assist functions in identifying the cell coordinates of values of interest.
	"""
	cell_container = []
	for col in sheet.columns:
		for cell in col:
			if cell.value is not None:
				cell_container.append(cell)
	return cell_container

def patient_identifier(cell_container: list, sheet) -> str:
	"""
	Returns the value of the patient identifier.
	"""
	cell_list = cell_string(cell_container, 'PATIENT', 0, 10)
	if len(cell_list) > 1:
		warnings.warn("Multiple matches to string 'PATIENT'.")
	row = list(sheet[cell_list[0].row])
	for cell in row:
		if cell.value is None:
			row.remove(cell)
	ID = row[1].value
	for word in row:
		print(word)
	return ID

def date_identifier(cell_container: list, date: str = 'date') -> int:
	"""
	Returns the year the patient was treated in which is important to know what the form of the excel sheet looks like. <date> is the value of the cell that is the row label for the 	fraction dates. Assumes dates in the sheet will be parsed as datetime.datetime object
	"""
	row_index = cell_string_row_index(cell_container,date)
	for cell in cell_container:
		if isinstance(cell.value,dt.datetime):
			return cell.value.year

def fraction_date_extractor(cell_container: list, date: str = 'date') -> list:
	"""
	Returns list of dates for each fraction patient was treated. Assumes dates in the sheet will be parsed as datetime.datetime object
	"""
	dates = []
	cell_row = []
	for cell in cell_container:
		if cell.row == cell_string_row_index(cell_container,date):
			cell_row.append(cell)
	for cell in cell_row:
		if isinstance(cell.value, dt.datetime):
			dates.append(cell.value)
	return dates

def number_of_fractions(cell_container: list, date: str = 'date') -> int:
	"""
	Returns the number of fractions in the sheet based on the number of dates
	"""
	return len(fraction_date_extractor(cell_container,date))

def cell_string(cell_container: list, value: str, minimum: int = 0, maximum = sys.maxsize) -> tuple:
	"""
	Returns a list of cells whose value contains <value>. String comparison is case-insensitive using casefold function and first attempts a substring match. If more than one match is found, then an exact match (case-insensitive) is attempted. If no match is found, then a warning is thrown and an empty list is returned. Minimum and maximum refer to row numbers and not index positions of cell_container.
	"""
	count = 0
	data_cell = []
	value_spliced = []
	splice_site = -1
	if isinstance(value,str):
		value = value.casefold()
		if conf.SPLICE in value:
			value_spliced = value.split(conf.SPLICE)
			splice_site = 1
		else:
			value_spliced.append(value)
			splice_site = 0
	else:
		raise TypeError(f"Value is not of type string. Value of splice_site is {splice_site}. This value should be either 0 or 1.")
	for cell in cell_container:
		if (cell.row < minimum) or (cell.row > maximum):
			continue
		if isinstance(cell.value, str):
			#if conf.DEBUG:
				#print(f'value to match is {value_spliced[splice_site]}')
			if value_spliced[splice_site] in cell.value.casefold():
				data_cell.append(cell)
				count+=1
	if count > 1:
		if conf.DEBUG:
			print(f"Ambiguous string matches: {data_cell}")
		for cell in data_cell:
			if cell.value.casefold() != value_spliced[splice_site].casefold():
				data_cell.remove(cell)

	if conf.DEBUG:
		if count > 1:
			warnings.warn(f"Search pattern <{value}> is ambiguous, {count} cells in the sheet  contain this pattern. An exact match was attempted. Remaining number of elements is {len(data_cell)}.")
	if len(data_cell) == 0:
		warnings.warn("Search pattern <" + value + "> was not found or failed exact match upon duplicate entry detection. Returned empty tuple.")
	return data_cell

def cell_string_row_index(cell_container: list, value: str) -> int:
	"""
	Returns the row number of the cell whose value is <value>. String comparison is case-insensitive using casefold function.
	"""
	_tuple = cell_string(cell_container,value)
	if len(_tuple) > 1:
		warnings.warn("Search pattern was ambiguous, row index for first match is returned")
	return _tuple[0].row

def cell_string_col_index(cell_container: list, value: str) -> int:
	"""
	Redundant but left for completeness. Returns the col index of the cell whose value is <value>. String comparison is case-insensitive using casefold function.
	"""
	_tuple = cell_string(cell_container,value)
	if len(_tuple) > 1:
		warnings.warn("Search pattern was ambiguous, col index for first match is returned")
	return _tuple[0].column

def data_extractor(cell_container: list, sheet, value: str, fraction: int):
	"""
	This is THE function to retrieve data given a string value to match. The data is extracted via coordinates rather than directly from cell_container for the purposes of clarity to the user.
	"""
	column_shift = fraction
	cell_of_interest = None
	# Get tuple of cells that match value
	matched_cells = []
	# If the name of a structure is contained in value, then find constraints
	if structure_check(value):
		constraints = duplicate_title_constraints(cell_container, value)
		if constraints is None:
			return None
		matched_cells = cell_string(cell_container, value, constraints[0], constraints[1])
		if len(matched_cells) > 1:
			print(f'{len(matched_cells)} were found. First match will be used. Please check values for <{value}>')
		#for cell in matched_cells:
			# Check which cell with duplicate title falls within the constraints
		if len(matched_cells) == 0:
			print(f'{value} could not be matched!')
			return None
		row = matched_cells[0].row
		if conf.DEBUG:
			print(row)
		cell_of_interest = matched_cells[0]
			# 
			# if row >= constraints[0] and row < constraints[1]:
			# 	cell_of_interest = cell

			# cell of interest is the structure header implies data of interest is the volume of the structure 
		if row == constraints[0]:
			column_shift += 1
	else:
		if conf.SPLICE in value:
			raise ValueError(f"Provided structure name in <{value}> was not found in conf.TITLES. Please double check entry or add structure to conf.TITLES")
		else:
			matched_cells = cell_string(cell_container, value)
			cell_of_interest = matched_cells[0]

	try:
		if ebrt_check(value.split(conf.SPLICE)[0]):
			column_shift = 1
	except TypeError as e:
		print (e)
	data_row = cell_of_interest.row
	data_column = get_column_letter(cell_of_interest.column + column_shift)
	
	# Get the data
	data = sheet[data_column + str(data_row)].value

	# Format the data to YMD if datetime
	# if isinstance(data, dt.datetime):
	# 	data = data.strftime("%m/%d/%Y") 

	return data

def pair_section_header(cell_container: list, value: str='EXTERNAL BEAM THERAPY') -> tuple:
	"""
	Returns the cells in the column containing the section headers.
	"""
	for idx, cell in enumerate(cell_container):
		if isinstance(cell.value, str) and get_column_letter(cell.column) == 'A':
			if value in cell.value:
				print(f'from pair_section_header: {cell_container[idx].value, cell_container[idx + 1].value}')
				return cell_container[idx], cell_container[idx + 1]
	return None

def duplicate_title_constraints(cell_container: list, value: str) -> list:
	"""
	If desired data has a duplicate title (e.g. D 100 exists for each structure), then we need to know the next section header after the one that is being requested to identify the desired value. This function returns a list of the rows for the section header of the desired data and the one directly after it. The returned data value must fall between these two rows.
	"""
	row_constraint = []
	# Variable for the structure name as it would appear in the excel sheet
	section_header = ''
	for key, list_pseudonyms in conf.TITLES.items():
		for title_value in list_pseudonyms:
			if title_value.casefold() in value.casefold():
				section_header = key
	# Get the cells of the section headers for the section of interest and the one after it
	section_header_pair = pair_section_header(cell_container, section_header)

	if section_header_pair is None:
		warnings.warn(f"{section_header} was not found.")
		return None

	max_constraint_cells = cell_string(cell_container,section_header_pair[1].value)

	min_constraint_row = cell_string_row_index(cell_container,section_header_pair[0].value)
	max_constraint_row = -999

	# Minimum row constraint: is the row number of the section that contains the desired data
	row_constraint.append(min_constraint_row)
	# Max row constraint: is the row number of the section following the section that contains the desired data.
	# In the event of finding duplicate section headers; check which duplicate has row > min row
	max_constraint_cells_row_indices = []
	for cell in max_constraint_cells:
		if cell.row < min_constraint_row:
			max_constraint_cells.remove(cell)
	row_difference = 999

	# Check for which remaining duplicate has the closest row number to the minimum row
	closest_cell = max_constraint_cells[0]
	for cell in max_constraint_cells:
		if (cell.row - min_constraint_row) < row_difference:
			row_difference = cell.row - min_constraint_row
			closest_cell = cell
	max_constraint_row = closest_cell.row
	row_constraint.append(max_constraint_row)
	if conf.DEBUG:
		print(f'DEBUG: from duplicate_title_constraints: {row_constraint}')
	return row_constraint

def structure_check(value: str, test_string: str='') -> bool:
	"""
	Check if value contains the name of a structure in TITLES.
	"""
	for key, list_pseudonyms in conf.TITLES.items():
		for title_value in list_pseudonyms:
			if title_value.casefold() in value.casefold():
				if conf.DEBUG:
					print(f'title_value is: {title_value}')
				return True

def ebrt_check(value: str) -> bool:
	"""
	Check if structure is EBRT. Used to not shift the column over one in each iteration over the fractions.
	"""
	for ebrt_pseudonyms in conf.TITLES['EXTERNAL BEAM THERAPY']:
		if conf.DEBUG:
				print(f'ebrt_check: value is {value}')
		if value.casefold() == ebrt_pseudonyms.casefold():
			return True
		else:
			return False


def increment_char(character: str,increment: int) -> str:
	"""
	Returns the next character in the alphabet e.g. A --> B
	"""
	return chr(ord(character) + increment)

def DVH_field_extractor(cell_container: list, sheet, match: str, shift: int=1):
	"""
	Returns the value of the cell with contents <match>. <shift> refers to how many cells over in the excel sheet the corresponding value is located.
	"""
	values = []
	value_cells = cell_string(cell_container,match)
	for cell in value_cells:
		cell_row = cell.row
		cell_col = get_column_letter(cell.column + shift)
		values.append(sheet[cell_col + str(cell_row)].value)
	return values
